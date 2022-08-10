'''
Description: 
Version: 1.0
Autor: z.cejay@gmail.com
Date: 2022-06-22 16:15:42
LastEditors: cejay
LastEditTime: 2022-08-10 09:19:34
'''

import json
import openpyxl


excelFile = "daoatlas.xlsx"

# read excel file and return a dictionary of sheets


def readExcel(fileName: str, sheetName: str = 'Sheet2', ignoreRow: int = 2):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb[sheetName]
    list = []
    _begin = False
    for row in sheet.iter_rows(min_row=ignoreRow, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if all(cell.value is None for cell in row):
            continue
        if _begin:
            list.append(row)

        if row[0].value == 'title':
            _begin = True

    data = []
    for i in range(len(list)):
        row = list[i]
        if row[0].value is None:
            # contributors
            pass

        else:
            # 新的DAO
            """ 
            title、tag、image、summary、logo、name、website、twitter、discord、forum、github
            """
            dao = {
                'title': row[0].value,
                'tag': row[1].value,
                'image': 'https://dao.hidev.cn/storage/images/'+row[2].value,
                'summary': row[3].value,
                'logo': 'https://dao.hidev.cn/storage/images/'+row[4].value,
                'name': row[5].value,
                'website': row[6].value,
                'twitter': row[7].value,
                'discord': row[8].value,
                'forum': row[9].value,
                'github': row[10].value,
                'contributorsArr': [],
                'contributors': '',
                'introduction': row[23].value,
                'GovernanceStructure_voting_snapshot': row[25].value,
                'GovernanceStructure_forum_discourse': list[i+1][25].value,
                'GovernanceStructure_content': list[i+2][25].value,
                'Tokenomics': row[26].value,
                'Contributor Incentives': row[27].value,
                'Howtojoin': row[28].value,
                'Editor': row[30].value,
                'Reviewer': row[31].value,
            }
            for i in range(0, 5):
                _img = row[12+i*2].value
                if _img is not None:
                    dao['contributorsArr'].append({
                        'key_pic': 'https://dao.hidev.cn/storage/images/' + row[12+i*2].value,
                        'key_contributors_name': list[i+1][12+i*2].value,
                        'key_contributors_job': list[i+2][12+i*2].value,
                        'key_contributors_intro': list[i+3][12+i*2].value,
                        'key_contributors_url': list[i+4][12+i*2].value,
                    })
            dao['contributors'] = json.dumps(dao['contributorsArr'])
            data.append(dao)

    return data


if __name__ == "__main__":
    newExcel = readExcel(excelFile)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAO Atlas"
    # write header
    ws.append(['title', 'tag', 'image', 'summary', 'logo', 'name', 'website', 'twitter', 'discord', 'forum', 'github', 'contributors', 'introduction', 'GovernanceStructure_voting_snapshot',
              'GovernanceStructure_forum_discourse', 'GovernanceStructure_content', 'Tokenomics', 'Contributor Incentives', 'Howtojoin', 'Editor', 'Reviewer'])
    # write data
    for dao in newExcel:
        ws.append([dao['title'], dao['tag'], dao['image'], dao['summary'], dao['logo'], dao['name'], dao['website'], dao['twitter'], dao['discord'], dao['forum'], dao['github'], dao['contributors'], dao['introduction'],
                  dao['GovernanceStructure_voting_snapshot'], dao['GovernanceStructure_forum_discourse'], dao['GovernanceStructure_content'], dao['Tokenomics'], dao['Contributor Incentives'], dao['Howtojoin'], dao['Editor'], dao['Reviewer']])

    wb.save(excelFile+"_new.xlsx")
